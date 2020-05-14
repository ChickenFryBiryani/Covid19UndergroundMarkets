import time
import datetime
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from mysql_cryptomarketsdb import MySQLcryptomarketsDB
from selenium_networksetting import *

# ++++++++++++++++++++ You need to update the following 3 variables ++++++++++++++++++++
# Student Name Abbr
g_sStudentNameAbbr = "og" # You need to update this
# Your Username and Password for Empire Market: http://empiremktxgjovhm.onion => sign up
g_sMarketUserName  = "canadaking"    # You need to update this
g_sMarketPassword  = "46357635"      # You need to update this
# You must register your own account in Empire Market.
# If two students use the same username, the website will show the error message:
# "The account is logged in somewhere else"
# This error message will prevent you from scraping further. So use your own account.
# Make sure to keep anonymous when you sign up. Don't use real identity!
# ++++++++++++++++++++ You need to update the above 3 variables ++++++++++++++++++++++++
# No need to change anything else below
# Scraping frequency for 4 kinds of web pages
g_nScrapingFreqDaysProductDesc = 30 # days
g_nScrapingFreqDaysProductRating = 7 # days
g_nScrapingFreqDaysVendorProfile = 30 # days

g_nScrapingFreqDaysVendorRating = 7 # days
# market information
g_sMarketNameAbbr = "ch"
# g_sMarketURL = ""
g_sMarketURL = "http://canadahqx53lcurj.onion"
g_nMarketGlobalID = 27
# local output data directory
g_sOutputDirectoryTemp = "/home/rob/covid19_ch/"
# time to wait for human to input CAPTCHA
g_nTimeSecToWait = 30 * 24 * 60 * 60  # 30 days
# set it True if you want to use the default username and password
g_bUseDefaultUsernamePasswd = False
# category for scraping. Examples: 'Drugs & Chemicals', 'Fraud'
#g_sCategoryName = 'Fraud' # You may update the category name if you need. Go to the website and find our other names.

# g_sCategoryToScrape = 'banklogs' #
# g_sCategoryToScrape = 'profiles' #
g_sCategoryToScrape = 'credit-cards' #
# g_sCategoryToScrape = 'accounts' #
# g_sCategoryToScrape = 'scampages' #
# g_sCategoryToScrape = 'dumps' #
# g_sCategoryToScrape = 'fraud-other' #
# g_sCategoryToScrape = 'ids' #
# g_sCategoryToScrape = 'orders' #
# g_sCategoryToScrape = 'counterfeits' #
# g_sCategoryToScrape = 'services-other' #


def Login(aBrowserDriver):
    global g_sMarketURL # needed to modify global copy of this variable; no need if you only read from it.
    aBrowserDriver.get(g_sMarketURL+'/login')
    #aInputElementCaptcha = aBrowserDriver.find_element_by_xpath("//input[contains(@name,'captcha')]")
    #aInputElementCaptcha.send_keys('')
    #aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
    #aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'Captcha entered successful. Please use the following link to access the market:')]")))
    #aNewURLOnionLink1 = aBrowserDriver.find_element_by_xpath("//div[contains(text(),'Captcha entered successful. Please use the following link to access the market:')]")
    #aNewURLOnionLink2 = aNewURLOnionLink1.find_element_by_xpath("//a[contains(@href,'http://')]")
    #g_sMarketURL = aNewURLOnionLink2.get_attribute('text')
    #aNewURLOnionLink2.click()

    ##aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
    # aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'The Canadian HeadQuarters')]")))
    time.sleep(3)
    # aLoginElement = aBrowserDriver.find_element_by_xpath("//a[contains(@href,'/login')]")
    # aLoginElement.click()
    # fill the username and password
    aInputElementUsername = aBrowserDriver.find_element_by_xpath("//input[contains(@name,'username')]")
    aInputElementUsername.send_keys(g_sMarketUserName)
    aInputElementPassword = aBrowserDriver.find_element_by_xpath("//input[contains(@name,'password')]")
    aInputElementPassword.send_keys(g_sMarketPassword)
    aInputElementCaptcha = aBrowserDriver.find_element_by_xpath("//input[contains(@name,'captcha')]")
    aInputElementCaptcha.send_keys('')
    time.sleep(5)
    aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
    sStringToBePresent = "//a[contains(@href,'/logout" + "')]"
    aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, sStringToBePresent)))
    # aProductsElement = aBrowserDriver.find_element_by_xpath("//a[contains(@href,'/products')]")
    # aProductsElement.click()
    # aCategoryElement = aBrowserDriver.find_element_by_xpath("//div[contains(@label,'+Fraud')]")
    # aCategoryElement.click()
    # aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait) # Wait up to x seconds (30 days).
    time.sleep(3)
    sStringOfCategry = g_sMarketURL+"/products/" + g_sCategoryToScrape
    aBrowserDriver.get(sStringOfCategry)
    # aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, sStringOfCategry)))
    # aCategoryElement = aBrowserDriver.find_element_by_xpath(sStringOfCategry)
    # aCategoryElement.click()

def NavigateToOnePage(aBrowserDriver, sPageLink):
    # "aBrowserDriver" : web driver
    # "sPageLink" : the link of the web page
    global g_sMarketURL # needed to modify global copy of this variable; no need if you only read from it.
    while True:
        aBrowserDriver.get(sPageLink)
        # Get the page title
        sPageTitle = aBrowserDriver.title
        if '(Privoxy@localhost)' in sPageTitle:
            # 502 - No server or forwarder data received (Privoxy@localhost)
            # 503 - Forwarding failure (Privoxy@localhost)
            # 504 - Connection timeout (Privoxy@localhost)
            aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
            aWaitNextPage.until(EC.element_to_be_clickable((By.LINK_TEXT, "Privoxy")))
        elif 'DDOS CAPTCHA' in sPageTitle:
            aInputElementCaptcha = aBrowserDriver.find_element_by_xpath("//input[contains(@name,'captcha')]")
            aInputElementCaptcha.send_keys('')
            aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
            aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'Captcha entered successful. Please use the following link to access the market:')]")))
            aNewURLOnionLink1 = aBrowserDriver.find_element_by_xpath("//div[contains(text(),'Captcha entered successful. Please use the following link to access the market:')]")
            aNewURLOnionLink2 = aNewURLOnionLink1.find_element_by_xpath("//a[contains(@href,'http://')]")
            g_sMarketURL = aNewURLOnionLink2.get_attribute('text')
            aNewURLOnionLink2.click()
        elif 'DDOS Protection' in sPageTitle:
            aWaitNextPage = WebDriverWait(aBrowserDriver, g_nTimeSecToWait)  # Wait up to x seconds (30 days).
            aWaitNextPage.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@href,'/login')]")))
            aLoginLink = aBrowserDriver.find_element_by_xpath("//a[contains(@href,'/login')]")
            aLoginLink.click()
        else:
            break
        # Sleep for a while and then go to next page
        time.sleep(4)
    # Wait for 1 second before scraping next page.
    time.sleep(4)

# The main function, the entry point
if __name__ == '__main__':
    # query the SQL server to retrieve some basic information

    aMySQLcrptmktDB = MySQLcryptomarketsDB()
    aMySQLcrptmktDB.m_sStudentNameAbbr = g_sStudentNameAbbr
    aMySQLcrptmktDB.m_sMarketNameAbbr = g_sMarketNameAbbr
    aMySQLcrptmktDB.m_nScrapingFreqDaysProductDesc   = g_nScrapingFreqDaysProductDesc
    aMySQLcrptmktDB.m_nScrapingFreqDaysProductRating = g_nScrapingFreqDaysProductRating
    aMySQLcrptmktDB.m_nScrapingFreqDaysVendorProfile = g_nScrapingFreqDaysVendorProfile
    aMySQLcrptmktDB.m_nScrapingFreqDaysVendorRating  = g_nScrapingFreqDaysVendorRating
    aMySQLcrptmktDB.MySQLQueryBasicInfor()
    g_nMarketGlobalID = aMySQLcrptmktDB.m_nMarketGlobalID
    if g_bUseDefaultUsernamePasswd == True:
        g_sMarketUserName = aMySQLcrptmktDB.m_sMarketUserName # this username and passwd are retrieved from DB server
        g_sMarketPassword = aMySQLcrptmktDB.m_sMarketPassword

    # Setup Firefox browser for visiting .onion sites through Tor (AWS proxy)
    aBrowserDriver = selenium_setup_firefox_network()
    #aBrowserDriver.set_window_rect(990, 640, 600, 660) # x,y,width,height; screen width 1853, height 1145

    # Visit the Darknet Market sites and Login
    Login(aBrowserDriver)

    # Get the total # of pages: "nMaxIndexOfPage"
    nIndexOfElements = 0 # the Index of the page nav element.
    nMaxIndexOfPage = 0  # The maximum number of page index, i.e., the total number of pages.
    # vElementPageNav = aBrowserDriver.find_element_by_xpath('//span[contains(text(),"1 of ")]').text
    # nMaxIndexOfPage = int(vElementPageNav[5:])
    vAllPageLinks = []
    pageList = aBrowserDriver.find_elements_by_class_name('pagination')
    # print(pageList)
    pageIndex = []
    if len(pageList) != 0:
        pageIndex = (pageList[0]).text
    if len(pageIndex) == 0:
        nMaxIndexOfPage = 1
    else:
        nMaxIndexOfPage = int(pageIndex[-4])*10 + int(pageIndex[-3])
    # print("#######")
    # print(pageIndex[-4])
    # print(pageIndex[-5])
    # print (pageIndex[-6])
    # print("#######")
    # print ((data1[-1]))
    # nMaxIndexOfPage = int((data[-2]).text)
    print("Total number of pages %d" % nMaxIndexOfPage)

    # Scrapes from this page
    nIdxOfPage = 1
    for i in range(nIdxOfPage, nMaxIndexOfPage+1):
        sOnePageLink = g_sMarketURL + "/products/" + g_sCategoryToScrape + "?sort=created_at&page=" + str(i)
        vAllPageLinks.append(sOnePageLink)
    print("Total number of pages %d" % nMaxIndexOfPage)
    #print(vAllPageLinks)

    # Start scraping. For each page, crawl all products in the page
    for sOnePageLink in vAllPageLinks: # nIdxOfPage is from 0 to (nMaxIndexOfPage - 1)
        #print(sOnePageLink)
        time.sleep(2)
        nPageIndex = nIdxOfPage
        nIdxOfPage += 1
        print("page %d" % nPageIndex)
        # Example page
        NavigateToOnePage(aBrowserDriver, sOnePageLink)
        # For each product in this page, visit its own page
        # First, get all the products on this page, insert them into the list of all products.
        vElementsProducts = aBrowserDriver.find_elements_by_class_name("product")
        # vElementsVendors = aBrowserDriver.find_elements_by_xpath("//a[contains(@href,'/user/')]")
        # You must call the above two lines first. In the following, we will leave this page.
        vAllProductsInThisPage = []
        for aElementOneProduct in vElementsProducts:
            sProductHref = aElementOneProduct.get_attribute("href")
            time.sleep(.2)
            if sProductHref not in vAllProductsInThisPage:
                vAllProductsInThisPage.append(sProductHref) # Need to remove duplicates
        #print(vAllProductsInThisPage)

        # Second, get all the vendors on this page, insert them into the list of all vendors.
        # vAllVendorsInThisPage = []
        # for aElementOneVendor in vElementsVendors:
        #     sVendorHref = aElementOneVendor.get_attribute("href")
        #     time.sleep(.2)
        #     if sVendorHref not in vAllVendorsInThisPage:
        #         vAllVendorsInThisPage.append(sVendorHref)  # Need to remove duplicates
        # print("%d products, %d vendors" % (len(vAllProductsInThisPage), len(vAllVendorsInThisPage)))
        print("%d products" % (len(vAllProductsInThisPage)))
        # Scrape the product description pages in "vAllProductsNew"
        NewProductCount = 0
        NewVendorCount = 0
        for sProductHref in vAllProductsInThisPage:
            time.sleep(1)
            sProductMarketID = sProductHref[sProductHref.rindex('/') + 1:]
            # 1. Scrape the product information
            # Example product page
            # bWhetherScraping, bNewProduct = aMySQLcrptmktDB.CheckWhetherScrapingProductDescription(sProductMarketID)
            bWhetherScraping = aMySQLcrptmktDB.CheckWhetherScrapingProductDescription(sProductMarketID) # OG change
            bNewProduct=True
            if bNewProduct == True:
                NewProductCount += 1
            # bWhetherScraping= True

            if bWhetherScraping == True:
                NavigateToOnePage(aBrowserDriver, sProductHref)
                # Save the html
                sCurrentUTCTime = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")
                aMySQLcrptmktDB.m_sCurrentUTCTime = sCurrentUTCTime
                sLocalOutputFileName = sCurrentUTCTime + '_' + str(g_nMarketGlobalID).zfill(2) + '_' + sProductMarketID + '_pd'
                sLocalOutputFileNameFullPath = g_sOutputDirectoryTemp + sLocalOutputFileName
                # Get screen shot of Product Description
                #aBrowserDriver.get_screenshot_as_file(sLocalOutputFileNameFullPath + "_img.png")
                aFile = open(sLocalOutputFileNameFullPath, "w")
                aFile.write(aBrowserDriver.page_source)
                aFile.close()
                aMySQLcrptmktDB.UpdateDatabaseUploadFileProductDescription(sLocalOutputFileName, sLocalOutputFileNameFullPath, sProductMarketID)
                os.remove(sLocalOutputFileNameFullPath)
                # Getting Vendor data
                vElementsVendorslist = aBrowserDriver.find_elements_by_xpath("//a[contains(@href,'/shop/')]")
                # print(vElementsVendorslist)
                # print(vElementsVendorslist.len())
                vElementsVendors = vElementsVendorslist[0].get_attribute("href")
                vElementsFeedbacklist = aBrowserDriver.find_elements_by_xpath("//a[contains(@href,'/feedback')]")
                vElementsFeedbacks = vElementsFeedbacklist[0].get_attribute("href")
                # print(vElementsVendors)
                sVendorHref = vElementsVendors
                sVendorMarketID = vElementsVendors[vElementsVendors.rindex('/') + 1:]
                if g_sMarketUserName == sVendorMarketID:
                    continue  # Skip the homepage of ourselves
                # 1. Scrape the vendor profile page
                # bWhetherScraping, bNewVendor = aMySQLcrptmktDB.CheckWhetherScrapingVendorProfile(sVendorMarketID)
                bWhetherScraping = aMySQLcrptmktDB.CheckWhetherScrapingVendorProfile(sVendorMarketID)

                bNewVendor=True
                if bNewVendor == True:
                    NewVendorCount += 1

                if bWhetherScraping == True:
                    sCurrentUTCTime = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")
                    aMySQLcrptmktDB.m_sCurrentUTCTime = sCurrentUTCTime
                    sLocalOutputFileName = sCurrentUTCTime + '_' + str(g_nMarketGlobalID).zfill(
                        2) + '_' + sVendorMarketID + '_vp'
                    sLocalOutputFileNameFullPath = g_sOutputDirectoryTemp + sLocalOutputFileName
                    NavigateToOnePage(aBrowserDriver, sVendorHref)
                    # Save the html
                    aFile = open(sLocalOutputFileNameFullPath, "w")
                    aFile.write(aBrowserDriver.page_source)
                    aFile.close()
                    # Move file to server, and update the sql database
                    aMySQLcrptmktDB.UpdateDatabaseUploadFileVendorProfile(sLocalOutputFileName,
                                                                          sLocalOutputFileNameFullPath, sVendorMarketID)
                    sCommandRemoveLocalfiles = 'rm ' + sLocalOutputFileNameFullPath
                    os.system(sCommandRemoveLocalfiles)
                    # Vendor rating
                    sLocalOutputFileName = sCurrentUTCTime + '_' + str(g_nMarketGlobalID).zfill(
                        2) + '_' + sVendorMarketID + '_vr'
                    NavigateToOnePage(aBrowserDriver, vElementsFeedbacks)
                    # Save the html
                    aFile = open(sLocalOutputFileNameFullPath, "w")
                    aFile.write(aBrowserDriver.page_source)
                    aFile.close()
                    # Move file to server, and update the sql database
                    aMySQLcrptmktDB.UpdateDatabaseUploadFileVendorProfile(sLocalOutputFileName,
                                                                          sLocalOutputFileNameFullPath, sVendorMarketID)
                    sCommandRemoveLocalfiles = 'rm ' + sLocalOutputFileNameFullPath
                    os.system(sCommandRemoveLocalfiles)

                bWhetherScraping = aMySQLcrptmktDB.CheckWhetherScrapingVendorRating(sVendorMarketID)
                if bWhetherScraping == True:
                    sVendorFeedbackHref = sVendorHref + "/feedback"
                    NavigateToOnePage(aBrowserDriver, sVendorFeedbackHref)
                    vaTablinkElements = aBrowserDriver.find_elements_by_class_name("tablinks ")
                    vsVendorRatingHRef = []
                    for aTablinkElement in vaTablinkElements:
                        sVendorRatingHref = aTablinkElement.get_attribute("href")
                        if 'p_feedback' in sVendorRatingHref or 'n_feedback' in sVendorRatingHref or 'c_feedback' in sVendorRatingHref or 'l_feedback' in sVendorRatingHref:
                            vsVendorRatingHRef.append(sVendorRatingHref)
                    # You have to extract all HRef link and then scrape them. Otherwise, error occurs.
                    sCurrentUTCTime = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")
                    aMySQLcrptmktDB.m_sCurrentUTCTime = sCurrentUTCTime
                    sLocalOutputFileName = sCurrentUTCTime + '_' + str(g_nMarketGlobalID).zfill(
                        2) + '_' + sVendorMarketID + '_vr'
                    sLocalOutputFileNameFullPath = g_sOutputDirectoryTemp + sLocalOutputFileName
                    for sVendorRatingHref in vsVendorRatingHRef:
                        # NavigateToOnePage(aBrowserDriver, sVendorRatingHref)
                        aFeedbackElement = aBrowserDriver.find_element_by_xpath("//a[contains(@href,'/feedback')]")
                        aFeedbackElement.click()
                        time.sleep(2)
                        # Save the html
                        sLocalOutputFileNameFullPath2 = ""
                        if 'p_feedback' in sVendorRatingHref:
                            sLocalOutputFileNameFullPath2 = sLocalOutputFileNameFullPath + '_p'
                        elif 'n_feedback' in sVendorRatingHref:
                            sLocalOutputFileNameFullPath2 = sLocalOutputFileNameFullPath + '_n'
                        elif 'c_feedback' in sVendorRatingHref:
                            sLocalOutputFileNameFullPath2 = sLocalOutputFileNameFullPath + '_c'
                        elif 'l_feedback' in sVendorRatingHref:
                            sLocalOutputFileNameFullPath2 = sLocalOutputFileNameFullPath + '_l'
                        aFile = open(sLocalOutputFileNameFullPath2, "w")
                        aFile.write(aBrowserDriver.page_source)
                        aFile.close()
                    sLocalOutputFileName3 = sLocalOutputFileName + '_q'  # 'q' represents question symbol
                    sLocalOutputFileNameFullPath3 = sLocalOutputFileNameFullPath + '_?'
                    aMySQLcrptmktDB.UpdateDatabaseUploadFileVendorRating(sLocalOutputFileName3,
                                                                         sLocalOutputFileNameFullPath3, sVendorMarketID)
                    sCommandRemoveLocalfiles = 'rm ' + sLocalOutputFileNameFullPath3
                    os.system(sCommandRemoveLocalfiles)
        print("New Products %d" % NewProductCount)
        #Saving new products to workbook
        # loc = ("/home/rob/Desktop/newautolist.xlsx")
        # wb = openpyxl.load_workbook("/home/rob/Desktop/newautolist.xlsx")
        # ws = wb.worksheets[0]
        # ws.cell(row=71, column=3).value = int(ws.cell(row=71, column=3).value) + NewProductCount
        # wb.save(loc)
        # No need to scrape the product rating page it is present in product description

        # Scrape the vendor profile in "vAllVendorsNew"
        # NewVendorCount = 0
        # for sVendorHref in vAllVendorsInThisPage:
        #     # Example vendor profile page
        #     sVendorMarketID = sVendorHref[sVendorHref.rindex('/') + 1:]
        #     if g_sMarketUserName == sVendorMarketID:
        #         continue # Skip the homepage of ourselves
        #     # 1. Scrape the vendor profile page
        #     bWhetherScraping, bNewVendor = aMySQLcrptmktDB.CheckWhetherScrapingVendorProfile(sVendorMarketID)
        #     #bWhetherScraping = aMySQLcrptmktDB.CheckWhetherScrapingVendorProfile(sVendorMarketID)
        #
        #     if bNewVendor == True:
        #         NewVendorCount += 1
        #
        #     if bWhetherScraping == True:
        #         sCurrentUTCTime = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")
        #         aMySQLcrptmktDB.m_sCurrentUTCTime = sCurrentUTCTime
        #         sLocalOutputFileName = sCurrentUTCTime + '_' + str(g_nMarketGlobalID).zfill(2) + '_' + sVendorMarketID + '_vp'
        #         sLocalOutputFileNameFullPath = g_sOutputDirectoryTemp + sLocalOutputFileName
        #         NavigateToOnePage(aBrowserDriver, sVendorHref)
        #         # Save the html
        #         aFile = open(sLocalOutputFileNameFullPath, "w")
        #         aFile.write(aBrowserDriver.page_source)
        #         aFile.close()
        #         # Move file to server, and update the sql database
        #         aMySQLcrptmktDB.UpdateDatabaseUploadFileVendorProfile(sLocalOutputFileName, sLocalOutputFileNameFullPath, sVendorMarketID)
        #         sCommandRemoveLocalfiles = 'rm ' + sLocalOutputFileNameFullPath
        #         os.system(sCommandRemoveLocalfiles)
        # print("New Vendors %d" % NewVendorCount)
        # # Saving new products to workbook
        # loc = ("/home/rob/Desktop/newautolist.xlsx")
        # wb = openpyxl.load_workbook("/home/rob/Desktop/newautolist.xlsx")
        # ws = wb.worksheets[0]
        # ws.cell(row=7, column=4).value = int(ws.cell(row=7, column=4).value) + NewVendorCount #Accounts
        # wb.save(loc)
    print("All jobs are done! Thanks for inputting so many CAPTCHAs!")
    aBrowserDriver.quit()
